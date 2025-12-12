# select_folder.py

import os
import openpyxl
import openpyxl.styles
import datetime as dt

def select_folder_path(request):

    folder_path = request.get('data', {}).get('project_folder', '')

    file_path = os.path.join(folder_path, '项目数据', '试算平衡表.xlsx')

    # 检查路径是否存在
    if os.path.exists(file_path):
        
        # 獲取設定值
        result_text = {
            'result_message': f'选择项目文件夹 {folder_path} 成功！\n'
        }

    else:
        
        # 创建目录及文件
        os.makedirs(os.path.join(folder_path, '项目数据'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '1.初步业务活动'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '2.风险评估'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '3.控制测试'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '4.实质性程序'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '5.其他项目'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '6.完成审计工作'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '7.永久性档案'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '8.底稿附件'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '8.底稿附件', '1.资产类资料'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '8.底稿附件', '2.负债类资料'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '8.底稿附件', '3.权益类资料'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '8.底稿附件', '4.损益类资料'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计底稿', '9.记账凭证检查拍照'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '审计报告'), exist_ok=True)
        os.makedirs(os.path.join(folder_path, '原始资料'), exist_ok=True)


        # 实例化工作簿
        wb = openpyxl.Workbook()

        # 获取第一张工作表并赋予它名称为基本信息
        ws = wb.active
        ws.title = '基本信息'

        ws['A1'].value = '被审计会计期间'
        ws['A2'].value = '被审计会计报表截止日'
        ws['A3'].value = '会计师事务所名称'
        ws['A4'].value = '保护单元格工作表密码'
        ws['A5'].value = '企业名称'
        ws['A6'].value = '成立日期'
        ws['A7'].value = '核准日期'
        ws['A8'].value = '统一社会信用代码'
        ws['A9'].value = '注册资本'
        ws['A10'].value = '法定代表人'
        ws['A11'].value = '注册地址'
        ws['A12'].value = '国标行业'
        ws['A13'].value = '登记机关'
        ws['A14'].value = '经营范围'

        #设置边框
        thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin'))

        for row in ws.iter_rows(min_row=1, min_col=1, max_row=14, max_col=2):
            for cell in row:
                cell.border = thin_border

        # 设置列宽、行高、自动换行
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.row_dimensions[14].height = 100
        ws['A14'].alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)
























        # 保存工作簿会在磁盘上创建文件
        wb.save(file_path)

        # # 使用 UTF-8 编码写入 JSON 文件
        # with open(settings_path, 'w') as f:
        #     json.dump(settings_dict, f, indent=4)

        result_text = {
            'result_message': f'创建项目文件夹 {folder_path} 的子文件夹成功！\n\n创建项目文件夹 {folder_path} 的试算平衡表文件成功！\n\n初始化项目文件夹 {folder_path} 成功！\n'
        }

    return ['select_folder_path', result_text]